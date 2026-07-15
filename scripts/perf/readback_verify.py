# -*- Python Version: 3.10 -*-

"""H2: Read key PHPP result cells to JSON, and compare two extracts.

Usage:
    # Extract from a SAVED workbook (openpyxl backend — default, no live Excel):
    python scripts/perf/readback_verify.py extract path/to/saved_phpp.xlsx -o out.json

    # Extract from the OPEN workbook (live xlwings block reads — manual only!):
    python scripts/perf/readback_verify.py extract --live -o out.json

    # Compare two extracts (exit code 1 on any mismatch):
    python scripts/perf/readback_verify.py compare a.json b.json --rtol 1e-6

The cell set is defined in a JSON spec file (default:
'readback_spec_en_v10.6.json' next to this script). Spec entry kinds:

    cell          {"name", "sheet", "address"}
    locator_cell  {"name", "sheet", "locator_col", "locator_string", "value_col",
                   "row_offset"=0, "row_start"=1, "row_end"=200, "match"="contains"}
                  -> find the label in a column, read 'value_col' at that row+offset.
    block_count   {"name", "sheet", "header_col", "header_string", "value_col",
                   "max_rows"=500, "stop_col"?, "stop_string"?, "numeric_only"=false,
                   "filter_col"?, "filter_kind"? ("numeric"|"contains"), "filter_value"?}
                  -> count non-empty (or numeric) cells in 'value_col' below a
                     section header, stopping at the next section header.
                     Omit header_col/header_string and give "start_row" instead
                     for fixed-position blocks. The optional filter keeps only
                     rows whose 'filter_col' cell is numeric / contains a string
                     — this is how user entry-rows are told apart from PHPP's
                     sub-header text and built-in library rows.
    block_sum     same fields as block_count -> sum of the numeric values.

The openpyxl backend reads cached formula VALUES from a saved file — these are
trustworthy because Excel computed them before the save. It never needs (or
touches) a live Excel application.
"""

import argparse
import json
import math
import pathlib
import sys
from typing import Any, Protocol

SCRIPT_DIR = pathlib.Path(__file__).resolve().parent
DEFAULT_SPEC = SCRIPT_DIR / "readback_spec_en_v10.6.json"

Value = Any  # str | float | int | bool | None


# -----------------------------------------------------------------------------
# -- Readers


class WorkbookReader(Protocol):
    """The two primitives the spec engine needs from any backend."""

    def read_cell(self, sheet: str, address: str) -> Value: ...

    def read_column(self, sheet: str, col: str, row_start: int, row_end: int) -> list[Value]: ...


class CachingReader:
    """Memoizes column reads so locator scans don't re-read the same column."""

    def __init__(self, reader: WorkbookReader):
        self.reader = reader
        self._column_cache: dict[tuple[str, str, int, int], list[Value]] = {}

    def read_cell(self, sheet: str, address: str) -> Value:
        return self.reader.read_cell(sheet, address)

    def read_column(self, sheet: str, col: str, row_start: int, row_end: int) -> list[Value]:
        key = (sheet, col, row_start, row_end)
        if key not in self._column_cache:
            self._column_cache[key] = self.reader.read_column(sheet, col, row_start, row_end)
        return self._column_cache[key]


class OpenpyxlReader:
    """Reads cached (post-calc) values from a SAVED .xlsx via openpyxl."""

    def __init__(self, path: pathlib.Path):
        import openpyxl

        self.wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def read_cell(self, sheet: str, address: str) -> Value:
        from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

        col_letter, row = coordinate_from_string(address)
        col = column_index_from_string(col_letter)
        ws = self.wb[sheet]
        for row_cells in ws.iter_rows(min_row=row, max_row=row, min_col=col, max_col=col):
            return row_cells[0].value
        return None

    def read_column(self, sheet: str, col: str, row_start: int, row_end: int) -> list[Value]:
        from openpyxl.utils.cell import column_index_from_string

        col_idx = column_index_from_string(col)
        ws = self.wb[sheet]
        values = [None] * (row_end - row_start + 1)
        for i, row_cells in enumerate(
            ws.iter_rows(min_row=row_start, max_row=row_end, min_col=col_idx, max_col=col_idx)
        ):
            values[i] = row_cells[0].value
        return values


class XlwingsReader:
    """Reads from the live open workbook via XLConnection block reads (manual only)."""

    def __init__(self, connection):
        self.connection = connection

    def read_cell(self, sheet: str, address: str) -> Value:
        return self.connection.get_single_data_item(sheet, address)

    def read_column(self, sheet: str, col: str, row_start: int, row_end: int) -> list[Value]:
        values = self.connection.get_single_column_data(sheet, col, row_start, row_end)
        if not isinstance(values, list):  # single-row range returns a scalar
            values = [values]
        return values


# -----------------------------------------------------------------------------
# -- Spec engine


def _matches(cell_value: Value, target: str, match: str) -> bool:
    if not isinstance(cell_value, str):
        return False
    if match == "exact":
        return cell_value.strip() == target
    return target.lower() in cell_value.lower()


def _find_row(
    reader: WorkbookReader, sheet: str, col: str, target: str, row_start: int, row_end: int, match: str
) -> int | None:
    values = reader.read_column(sheet, col, row_start, row_end)
    for i, value in enumerate(values):
        if _matches(value, target, match):
            return row_start + i
    return None


def _as_number(value: Value) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        try:
            number = float(str(value).strip())
        except (ValueError, TypeError):
            return None
    # -- NaN/inf are non-numbers here: a NaN would poison 'block_sum' totals,
    # -- and NaN != NaN would make 'compare()' flag two equivalent extracts.
    return number if math.isfinite(number) else None


def _block_values(reader: WorkbookReader, entry: dict) -> list[Value]:
    """Resolve a block_count/block_sum entry to the list of value-column cells."""
    sheet = entry["sheet"]
    max_rows = entry.get("max_rows", 500)

    if "start_row" in entry:
        first_row = entry["start_row"]
    else:
        header_row = _find_row(
            reader,
            sheet,
            entry["header_col"],
            entry["header_string"],
            entry.get("row_start", 1),
            entry.get("row_end", 500),
            entry.get("match", "contains"),
        )
        if header_row is None:
            raise LookupError(f"Header '{entry['header_string']}' not found in {sheet}!{entry['header_col']}")
        first_row = header_row + 1

    last_row = first_row + max_rows - 1
    values = reader.read_column(sheet, entry["value_col"], first_row, last_row)

    if stop_string := entry.get("stop_string"):
        stop_col = entry.get("stop_col", entry.get("header_col", entry["value_col"]))
        stop_values = reader.read_column(sheet, stop_col, first_row, last_row)
        match = entry.get("match", "contains")
        for i, stop_value in enumerate(stop_values):
            if _matches(stop_value, stop_string, match):
                values = values[:i]
                break

    if filter_col := entry.get("filter_col"):
        filter_values = reader.read_column(sheet, filter_col, first_row, last_row)
        if entry.get("filter_kind", "numeric") == "numeric":
            keep = [_as_number(v) is not None for v in filter_values]
        else:
            keep = [isinstance(v, str) and entry["filter_value"] in v for v in filter_values]
        values = [v for v, k in zip(values, keep) if k]
    return values


def _extract_entry(reader: WorkbookReader, entry: dict) -> Value:
    kind = entry["kind"]

    if kind == "cell":
        return reader.read_cell(entry["sheet"], entry["address"])

    if kind == "locator_cell":
        row = _find_row(
            reader,
            entry["sheet"],
            entry["locator_col"],
            entry["locator_string"],
            entry.get("row_start", 1),
            entry.get("row_end", 200),
            entry.get("match", "contains"),
        )
        if row is None:
            raise LookupError(
                f"Locator '{entry['locator_string']}' not found in {entry['sheet']}!{entry['locator_col']}"
            )
        return reader.read_cell(entry["sheet"], f"{entry['value_col']}{row + entry.get('row_offset', 0)}")

    if kind in ("block_count", "block_sum"):
        values = _block_values(reader, entry)
        if entry.get("numeric_only") or kind == "block_sum":
            numbers = [n for n in (_as_number(v) for v in values) if n is not None]
            return sum(numbers) if kind == "block_sum" else len(numbers)
        return sum(1 for v in values if v not in (None, ""))

    raise ValueError(f"Unknown spec entry kind: '{kind}'")


def extract(reader: WorkbookReader, spec: list[dict]) -> dict[str, Any]:
    """Run every spec entry against the workbook; errors become '#ERROR: ...' strings."""
    reader = CachingReader(reader)
    results: dict[str, Any] = {}
    for entry in spec:
        try:
            results[entry["name"]] = _extract_entry(reader, entry)
        except Exception as e:
            results[entry["name"]] = f"#ERROR: {e}"
    return results


# -----------------------------------------------------------------------------
# -- Compare


def compare(a: dict[str, Any], b: dict[str, Any], rtol: float = 1e-6, atol: float = 1e-9) -> list[str]:
    """Return a list of mismatch descriptions (empty list = extracts are equivalent)."""
    diffs: list[str] = []
    for key in sorted(set(a) | set(b)):
        if key not in a:
            diffs.append(f"'{key}': missing from first extract")
            continue
        if key not in b:
            diffs.append(f"'{key}': missing from second extract")
            continue
        va, vb = a[key], b[key]
        # -- NaN floats can round-trip through the JSON extracts; NaN != NaN,
        # -- so compare them by NaN-ness instead of equality.
        va_nan = isinstance(va, float) and math.isnan(va)
        vb_nan = isinstance(vb, float) and math.isnan(vb)
        if va_nan or vb_nan:
            if not (va_nan and vb_nan):
                diffs.append(f"'{key}': {va!r} != {vb!r}")
            continue
        na, nb = _as_number(va), _as_number(vb)
        if na is not None and nb is not None:
            if not math.isclose(na, nb, rel_tol=rtol, abs_tol=atol):
                diffs.append(f"'{key}': {va!r} != {vb!r}")
        elif va != vb:
            diffs.append(f"'{key}': {va!r} != {vb!r}")
    return diffs


# -----------------------------------------------------------------------------
# -- CLI


def _load_spec(path: pathlib.Path) -> list[dict]:
    return json.loads(path.read_text())["entries"]


def _cmd_extract(args) -> int:
    spec = _load_spec(pathlib.Path(args.spec))

    if args.live:
        import perf_paths

        perf_paths.confirm_live_excel_run(args.yes)
        import xlwings as xw

        from PHX.xl.xl_app import XLConnection

        workbook_path = pathlib.Path(args.workbook).resolve() if args.workbook else None
        connection = XLConnection(xl_framework=xw, xl_file_path=workbook_path)
        reader: WorkbookReader = XlwingsReader(connection)
        source = str(connection.wb.fullname)
    else:
        if not args.workbook:
            sys.exit("Error: a saved workbook path is required (or use --live).")
        workbook_path = pathlib.Path(args.workbook).resolve()
        if not workbook_path.exists():
            sys.exit(f"Error: workbook not found: {workbook_path}")
        reader = OpenpyxlReader(workbook_path)
        source = str(workbook_path)

    results = extract(reader, spec)
    payload = {"source": source, "spec": str(args.spec), "values": results}
    output = json.dumps(payload, indent=2, default=str)
    if args.out:
        pathlib.Path(args.out).write_text(output)
        print(f"Wrote {len(results)} values -> {args.out}")
    else:
        print(output)

    errors = [k for k, v in results.items() if isinstance(v, str) and v.startswith("#ERROR")]
    if errors:
        print(f"\nWARNING: {len(errors)} entries failed: {', '.join(errors)}", file=sys.stderr)
        return 2
    return 0


def _cmd_compare(args) -> int:
    a = json.loads(pathlib.Path(args.file_a).read_text())["values"]
    b = json.loads(pathlib.Path(args.file_b).read_text())["values"]
    diffs = compare(a, b, rtol=args.rtol, atol=args.atol)
    if diffs:
        print(f"FAIL: {len(diffs)} mismatches:")
        for diff in diffs:
            print(f"  - {diff}")
        return 1
    print(f"PASS: all {len(a)} values match (rtol={args.rtol}, atol={args.atol}).")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)

    p_extract = sub.add_parser("extract", help="Extract key cells from a PHPP workbook to JSON.")
    p_extract.add_argument("workbook", nargs="?", help="Path to a SAVED .xlsx (omit only with --live).")
    p_extract.add_argument("--spec", default=str(DEFAULT_SPEC), help="Cell-spec JSON file.")
    p_extract.add_argument("-o", "--out", help="Output JSON path (default: stdout).")
    p_extract.add_argument("--live", action="store_true", help="Read via live xlwings instead of openpyxl.")
    p_extract.add_argument("--yes", action="store_true", help="Skip the live-Excel confirmation prompt.")
    p_extract.set_defaults(func=_cmd_extract)

    p_compare = sub.add_parser("compare", help="Diff two extract JSONs.")
    p_compare.add_argument("file_a")
    p_compare.add_argument("file_b")
    p_compare.add_argument("--rtol", type=float, default=1e-6)
    p_compare.add_argument("--atol", type=float, default=1e-9)
    p_compare.set_defaults(func=_cmd_compare)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
